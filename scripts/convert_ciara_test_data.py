#!/usr/bin/env python3
"""Convert Ciara test data from PDF/XLSX to JPG/JSON format.

ENVIRONMENT: Run in weather-doc-extractor conda environment:
  conda activate weather-doc-extractor
  python scripts/convert_ciara_test_data.py

Input:
  - test_data/from_Ciara/originals/*.pdf → images
  - test_data/from_Ciara/originals/*.xlsx → transcriptions

Output:
  - test_data/from_Ciara/images/*.jpg
  - test_data/from_Ciara/transcriptions/*.json

XLSX layout:
  - Columns H-S: 12 months of daily data
  - Rows 2-32: Days 1-31
  - Column U rows 2-13: Monthly totals
"""

from __future__ import annotations

import json
import logging
from pathlib import Path

import openpyxl
from pdf2image import convert_from_path
from PIL import Image

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

IMAGES_OUTPUT = Path("test_data/from_Ciara/images")
TRANSCRIPTIONS_OUTPUT = Path("test_data/from_Ciara/transcriptions")
ORIGINALS_INPUT = Path("test_data/from_Ciara/originals")

# XLSX column letters for months H-S
MONTH_COLUMNS = ["H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S"]
DAY_ROWS = list(range(2, 33))  # Rows 2-32 for Days 1-31
TOTALS_COLUMN = "U"
TOTALS_ROWS = list(range(2, 14))  # Rows 2-13 for monthly totals


def to_drain_stem(source_stem: str) -> str:
    """Convert numeric Ciara stems to DRain-compatible stems."""
    if source_stem.startswith("DRain_"):
        return source_stem
    return f"DRain_1901-1910_Ciara-{source_stem}"


def convert_pdf_to_jpg(pdf_path: Path, output_path: Path) -> bool:
    """Convert PDF to JPG, resizing to portrait format."""
    try:
        logger.info(f"Converting {pdf_path.name} to JPG...")
        images = convert_from_path(str(pdf_path), first_page=1, last_page=1)
        if not images:
            logger.error(f"No pages extracted from {pdf_path}")
            return False

        img = images[0]
        # Resize to standard portrait (maintain aspect ratio, ~1020px height)
        # Target: 765x1020 for standard document
        img_resized = img.resize((765, 1020), Image.Resampling.LANCZOS)
        img_resized.save(output_path, "JPEG", quality=95)
        logger.info(f"  Saved: {output_path}")
        return True
    except Exception as e:
        logger.error(f"Failed to convert {pdf_path}: {e}")
        return False


def read_excel_value(cell) -> str:
    """Convert Excel cell value to string, handling None and numeric values."""
    if cell.value is None:
        return "null"
    val = cell.value
    if isinstance(val, (int, float)):
        # Format as string with reasonable precision
        if isinstance(val, float):
            return f"{val:.2f}".rstrip("0").rstrip(".")
        return str(val)
    s = str(val).strip()
    return "null" if s.lower() == "null" or s == "" else s


def excel_to_json(xlsx_path: Path, output_path: Path) -> bool:
    """Extract rainfall data from XLSX and convert to JSON format."""
    try:
        logger.info(f"Parsing {xlsx_path.name}...")
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active

        # Extract daily data (Days 1-31 from rows 2-32, columns H-S)
        grid: dict[str, list[str]] = {}
        for day_idx, row_num in enumerate(DAY_ROWS, start=1):
            day_key = f"Day {day_idx}"
            grid[day_key] = []
            for col_letter in MONTH_COLUMNS:
                cell = ws[f"{col_letter}{row_num}"]
                value = read_excel_value(cell)
                grid[day_key].append(value)

        # Extract totals (column U, rows 2-13 for months 1-12)
        totals = []
        for row_num in TOTALS_ROWS:
            cell = ws[f"{TOTALS_COLUMN}{row_num}"]
            value = read_excel_value(cell)
            totals.append(value)

        grid["Totals"] = totals

        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(json.dumps(grid, indent=2))
        logger.info(f"  Saved: {output_path}")
        return True
    except Exception as e:
        logger.error(f"Failed to convert {xlsx_path}: {e}")
        return False


def main() -> int:
    """Convert all Ciara test data files."""
    IMAGES_OUTPUT.mkdir(parents=True, exist_ok=True)
    TRANSCRIPTIONS_OUTPUT.mkdir(parents=True, exist_ok=True)

    if not ORIGINALS_INPUT.exists():
        logger.error(f"Input directory not found: {ORIGINALS_INPUT}")
        return 1

    # Get all PDF and XLSX files
    pdf_files = sorted(ORIGINALS_INPUT.glob("*.pdf"))
    xlsx_files = sorted(ORIGINALS_INPUT.glob("*.xlsx"))

    if not pdf_files:
        logger.warning(f"No PDF files found in {ORIGINALS_INPUT}")
    if not xlsx_files:
        logger.warning(f"No XLSX files found in {ORIGINALS_INPUT}")

    success_count = 0
    fail_count = 0

    # Convert images
    for pdf_path in pdf_files:
        stem = to_drain_stem(pdf_path.stem)
        output_path = IMAGES_OUTPUT / f"{stem}.jpg"
        if convert_pdf_to_jpg(pdf_path, output_path):
            success_count += 1
        else:
            fail_count += 1

    # Convert transcriptions
    for xlsx_path in xlsx_files:
        stem = to_drain_stem(xlsx_path.stem)
        output_path = TRANSCRIPTIONS_OUTPUT / f"{stem}.json"
        if excel_to_json(xlsx_path, output_path):
            success_count += 1
        else:
            fail_count += 1

    logger.info(
        f"\nConversion complete: {success_count} succeeded, {fail_count} failed"
    )
    return 0 if fail_count == 0 else 1


if __name__ == "__main__":
    import sys

    sys.exit(main())
